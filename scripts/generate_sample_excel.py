"""Genera un Excel de ejemplo con la estructura real para pruebas."""
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Instalá openpyxl: pip install openpyxl")
    sys.exit(1)

HEADERS = [
    "Material",
    "Material MLFB",
    "Stock Actual",
    "Stock En Viaje",
    "Stock Pendiente",
    "Stock a Comprar",
    "Costo Estante (USD / UN)",
    "Stock Actual x Costo(USD)",
]

ROWS = [
    ["786518 - Int Tmag 06kA C 2P 16A", "5TJ6216-7", 5947, 0, 0, 0, 4.08, 24263.76],
    ["786520 - Int Tmag 06kA C 2P 20A", "5TJ6220-7", 1200, 200, 100, 0, 4.50, 5400.00],
    ["790100 - Contactor 3P 9A 24VDC",  "3RT2016-2BB41", 45, 0, 20, 50, 38.50, 1732.50],
    ["790200 - Relé Térmico 6-9A",       "3RU2116-1JB0", 0, 0, 0, 30, "-", "-"],
    ["800015 - Cable H07V-K 1x1.5mm2",  "3G2.5-BL", 850, 150, 0, 0, 0.95, 807.50],
    ["810030 - Bornera 4mm2 Gris",       "8WH2000-0AA01", 3200, 0, 0, 0, 0.42, 1344.00],
    ["820010 - Fusible NH 00 63A",       "3NA3820", 120, 0, 50, 100, 7.20, 864.00],
    ["830005 - Pulsador Verde 22mm",     "3SB1400-0AA41", 60, 0, 0, 20, 12.30, 738.00],
]

def main():
    out = Path(__file__).parent.parent / "data" / "stock_ejemplo.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")

    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = max(len(header) + 4, 14)

    for row_data in ROWS:
        ws.append(row_data)

    wb.save(out)
    print(f"Excel generado: {out}")

if __name__ == "__main__":
    main()
