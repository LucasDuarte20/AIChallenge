import logging
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from sqlalchemy.orm import Session

from app.core.config import settings
from app.repositories.plan_repo import list_visits

logger = logging.getLogger(__name__)


def export_plan_to_xlsx(db: Session, executive_code: str, year: int) -> Path:
    visits = list_visits(db, executive_code=executive_code, year=year)
    out_dir = Path(settings.data_dir) / "exports"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"agenda_{executive_code}_{year}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Agenda {year}"

    headers = ["Fecha", "Hora", "Orden", "Cliente", "Dirección", "Departamento"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = max(len(h) + 6, 16)

    for i, v in enumerate(visits, start=2):
        # visit_date es datetime tz-aware; export como YYYY-MM-DD
        date_str = v.visit_date.date().isoformat() if v.visit_date else ""
        time_str = v.visit_date.time().strftime("%H:%M") if v.visit_date else ""
        ws.cell(row=i, column=1, value=date_str)
        ws.cell(row=i, column=2, value=time_str)
        ws.cell(row=i, column=3, value=v.day_order or "")
        ws.cell(row=i, column=4, value=v.customer_name or "")
        ws.cell(row=i, column=5, value=v.customer_address or "")
        ws.cell(row=i, column=6, value=v.department or "")

    wb.save(out_path)
    logger.info("Agenda exportada: %s", out_path)
    return out_path

