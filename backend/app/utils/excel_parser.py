import logging
from pathlib import Path
from typing import Any

import openpyxl

logger = logging.getLogger(__name__)

COLUMN_MAP: dict[str, str] = {
    "Material": "material_raw",
    "Material MLFB": "material_mlfb",
    "Stock Actual": "stock_actual",
    "Stock En Viaje": "stock_en_viaje",
    "Stock Pendiente": "stock_pendiente",
    "Stock a Comprar": "stock_a_comprar",
    "Costo Estante (USD / UN)": "costo_estante_usd_unit",
    "Stock Actual x Costo(USD)": "stock_actual_x_costo_usd",
}

NUMERIC_FIELDS = {
    "stock_actual",
    "stock_en_viaje",
    "stock_pendiente",
    "stock_a_comprar",
    "costo_estante_usd_unit",
    "stock_actual_x_costo_usd",
}


def _clean_value(value: Any, field: str) -> Any:
    if value is None or str(value).strip() in ("", "-", "–", "—", "N/A", "n/a"):
        return None
    if field in NUMERIC_FIELDS:
        try:
            return float(value)
        except (ValueError, TypeError):
            return None
    return str(value).strip()


def _split_material(raw: str) -> tuple[str | None, str | None]:
    """Extrae material_code y material_name del campo Material."""
    raw = raw.strip()
    if " - " in raw:
        code, name = raw.split(" - ", 1)
        return code.strip() or None, name.strip() or None
    return raw or None, None


def parse_excel(file_path: str | Path, sheet_name: str = "Sheet1") -> list[dict]:
    """
    Parsea el Excel y retorna lista de dicts listos para insertar en DB.
    Lanza ValueError si no encuentra la hoja o las columnas esperadas.
    """
    wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=True)

    if sheet_name not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        raise ValueError(f"Hoja '{sheet_name}' no encontrada. Disponibles: {available}")

    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)

    headers_raw = next(rows_iter, None)
    if headers_raw is None:
        raise ValueError("El archivo está vacío.")

    headers = [str(h).strip() if h is not None else "" for h in headers_raw]
    logger.info("Columnas encontradas: %s", headers)

    missing = [col for col in COLUMN_MAP if col not in headers]
    if missing:
        logger.warning("Columnas no encontradas (se ignorarán): %s", missing)

    results: list[dict] = []
    for row_idx, row in enumerate(rows_iter, start=2):
        if not any(v is not None for v in row):
            continue

        row_dict: dict[str, Any] = {}
        for col_name, field_name in COLUMN_MAP.items():
            if col_name in headers:
                idx = headers.index(col_name)
                raw_val = row[idx] if idx < len(row) else None
                row_dict[field_name] = _clean_value(raw_val, field_name)

        raw_material = row_dict.get("material_raw")
        if not raw_material:
            logger.debug("Fila %d ignorada: sin valor en columna Material", row_idx)
            continue

        code, name = _split_material(str(raw_material))
        row_dict["material_code"] = code
        row_dict["material_name"] = name

        results.append(row_dict)

    wb.close()
    logger.info("Total filas parseadas: %d", len(results))
    return results
