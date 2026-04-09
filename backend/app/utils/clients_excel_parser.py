import logging
import re
from pathlib import Path
from typing import Any

import openpyxl

logger = logging.getLogger(__name__)

CLIENTS_SHEET_DEFAULT = "LISTA DE INDUSTRIA"
# Si no encontramos encabezados, fallback (archivos viejos usaban fila 3; los nuevos suelen usar fila 2)
HEADER_ROW_FALLBACK = 3

# (campo_db, posibles nombres exactos de columna en el Excel)
FIELD_ALIASES: list[tuple[str, tuple[str, ...]]] = [
    ("customer_number", ("N. Cli.",)),
    ("name", ("Nombre", "1. Nombre")),
    ("department", ("Departamento",)),
    ("address_raw", ("Dirección",)),
    ("client_size", ("Tipo",)),
    ("potential", ("Potencial",)),
    ("phone", ("Teléfono",)),
    ("office", ("Oficina",)),
    ("executive_code", ("Vendedor",)),
    ("sector", ("Sector",)),
]


def _clean(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "-", "–", "—"):
        return None
    return s


def _normalize_size(s: str | None) -> str | None:
    if not s:
        return None
    key = s.lower().strip()
    key = re.sub(r"\s+", " ", key)
    mapping = {"chico": "chico", "medio": "medio", "grande": "grande"}
    return mapping.get(key, key)


def _normalize_potential(s: str | None) -> str | None:
    if not s:
        return None
    key = s.lower().strip()
    mapping = {"bajo": "bajo", "medio": "medio", "alto": "alto"}
    return mapping.get(key, key)


def find_header_row(ws, max_scan: int = 25) -> int:
    """
    Busca la fila donde están los encabezados (debe haber Dirección y Nombre).
    """
    for row_idx in range(1, max_scan + 1):
        row = next(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))
        vals = [(_clean(v) or "") for v in row]
        has_dir = "Dirección" in vals
        has_name = "Nombre" in vals or "1. Nombre" in vals
        if has_dir and has_name:
            logger.info("Encabezados detectados en fila %d", row_idx)
            return row_idx
    return HEADER_ROW_FALLBACK


def parse_clients_excel(
    file_path: str | Path,
    sheet_name: str = CLIENTS_SHEET_DEFAULT,
    header_row_index: int | None = None,
) -> list[dict]:
    """
    header_row_index:
      - None o 0: auto-detectar fila de encabezados
      - >= 1: usar esa fila (1 = primera fila del Excel)
    """
    # read_only=False: permite detectar fila de encabezados y luego leer todo el archivo.
    wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=False)
    if sheet_name not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        raise ValueError(f"Hoja '{sheet_name}' no encontrada. Disponibles: {available}")

    ws = wb[sheet_name]

    if header_row_index is None or header_row_index == 0:
        header_row = find_header_row(ws)
    else:
        header_row = header_row_index

    headers: list[str] | None = None
    results: list[dict] = []

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        vals = [(_clean(v) or "") for v in row]

        if row_idx == header_row:
            headers = vals
            logger.info("Headers clientes (fila %d): %s", header_row, headers)
            continue

        if headers is None:
            continue

        if not any(vals):
            continue

        item: dict[str, Any] = {}
        for field_name, aliases in FIELD_ALIASES:
            val: str | None = None
            for alias in aliases:
                if alias in headers:
                    idx = headers.index(alias)
                    val = _clean(row[idx] if idx < len(row) else None)
                    break
            item[field_name] = val

        item["client_size"] = _normalize_size(item.get("client_size"))
        item["potential"] = _normalize_potential(item.get("potential"))

        if not item.get("name"):
            continue

        results.append(item)

    wb.close()
    logger.info("Total clientes parseados: %d", len(results))
    return results
