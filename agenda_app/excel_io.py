from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl

from agenda_app.models import ClientRow

HEADER_ROW_FALLBACK = 3

FIELD_ALIASES: list[tuple[str, tuple[str, ...]]] = [
    ("customer_number", ("N. Cli.",)),
    ("name", ("Nombre", "1. Nombre")),
    ("department", ("Departamento",)),
    ("address_raw", ("Dirección",)),
    ("client_size", ("Tipo",)),
    ("potential", ("Potencial",)),
    ("phone", ("Teléfono",)),
    ("office", ("Oficina",)),
    ("executive_code", ("Vendedor",)),
    ("sector", ("Sector",)),
    ("last_visit", ("Última visita", "Fecha última visita", "Ultima visita", "Ultima visita-contacto")),
    ("next_visit", ("Próxima visita", "Proxima visita", "Fecha próxima visita", "Proxima visita-contacto", "Siguiente visita-contacto")),
    ("lat", ("Latitud", "Lat")),
    ("lon", ("Longitud", "Lon", "Lng")),
]


def _clean(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "-", "–", "—"):
        return None
    return s


def _normalize_size(s: str | None) -> str | None:
    if not s:
        return None
    key = re.sub(r"\s+", " ", s.lower().strip())
    mapping = {"chico": "chico", "medio": "medio", "grande": "grande"}
    return mapping.get(key, key)


def _normalize_potential(s: str | None) -> str | None:
    if not s:
        return None
    key = s.lower().strip()
    mapping = {"bajo": "bajo", "medio": "medio", "alto": "alto"}
    return mapping.get(key, key)


def _parse_date_cell(v: Any) -> date | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = _clean(v)
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    try:
        num = float(s)
        from openpyxl.utils.datetime import from_excel

        dt = from_excel(num)
        if isinstance(dt, datetime):
            return dt.date()
        if isinstance(dt, date):
            return dt
    except (ValueError, TypeError, ImportError):
        pass
    return None


def _parse_float(v: Any) -> float | None:
    if v is None:
        return None
    try:
        return float(str(v).replace(",", "."))
    except (ValueError, TypeError):
        return None


def find_header_row(ws, max_scan: int = 25) -> int:
    for row_idx in range(1, max_scan + 1):
        row = next(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))
        vals = [(_clean(v) or "").lower() for v in row]
        has_dir = "dirección" in vals or "direccion" in vals
        has_name = "nombre" in vals or "1. nombre" in vals
        if has_dir and has_name:
            return row_idx
    return HEADER_ROW_FALLBACK


CLIENTS_SHEET_DEFAULT = None

def load_clients(
    file_path: str | Path,
    sheet_name: str | None = CLIENTS_SHEET_DEFAULT,
    header_row_index: int | None = None,
) -> tuple[list[ClientRow], openpyxl.workbook.workbook.Workbook, str, int]:
    """
    Retorna clientes, workbook abierto (para guardar después), nombre de hoja, fila de encabezado.
    El llamador debe cerrar el workbook si no se usa save_workbook.
    """
    path = Path(file_path)
    wb = openpyxl.load_workbook(str(path), data_only=False)
    
    if not sheet_name:
        sheet_name = wb.sheetnames[0]
    elif sheet_name not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        raise ValueError(f"Hoja '{sheet_name}' no encontrada. Disponibles: {available}")

    ws = wb[sheet_name]
    if header_row_index is None or header_row_index == 0:
        header_row = find_header_row(ws)
    else:
        header_row = header_row_index

    headers: list[str] | None = None
    results: list[ClientRow] = []

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        vals = [(_clean(v) or "") for v in row]

        if row_idx == header_row:
            headers = vals
            headers_lower = [h.lower() for h in headers]
            continue

        if headers is None:
            continue

        if not any(vals):
            continue

        item: dict[str, Any] = {}
        for field_name, aliases in FIELD_ALIASES:
            val: str | None = None
            for alias in aliases:
                if alias.lower() in headers_lower:
                    idx = headers_lower.index(alias.lower())
                    raw = row[idx] if idx < len(row) else None
                    if field_name in ("last_visit", "next_visit"):
                        item[field_name] = _parse_date_cell(raw)
                    elif field_name in ("lat", "lon"):
                        item[field_name] = _parse_float(raw)
                    else:
                        val = _clean(raw)
                        item[field_name] = val
                    break
            if field_name not in item:
                item[field_name] = None

        item["client_size"] = _normalize_size(item.get("client_size"))
        item["potential"] = _normalize_potential(item.get("potential"))

        name = item.get("name")
        if not name:
            continue

        results.append(
            ClientRow(
                row_index=row_idx,
                customer_number=item.get("customer_number"),
                name=name,
                department=item.get("department"),
                address_raw=item.get("address_raw"),
                client_size=item.get("client_size"),
                potential=item.get("potential"),
                phone=item.get("phone"),
                office=item.get("office"),
                executive_code=item.get("executive_code"),
                sector=item.get("sector"),
                last_visit=item.get("last_visit"),
                next_visit=item.get("next_visit"),
                lat=item.get("lat"),
                lon=item.get("lon"),
            )
        )

    return results, wb, sheet_name, header_row


def _header_row_values(ws, header_row: int) -> list[str]:
    row = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    return [(_clean(v) or "") for v in row]


def _col_letter_for_field(headers: list[str], field: str) -> str | None:
    _, aliases = next((fa for fa in FIELD_ALIASES if fa[0] == field), ("", ()))
    headers_lower = [h.lower() for h in headers]
    for a in aliases:
        if a.lower() in headers_lower:
            idx = headers_lower.index(a.lower()) + 1
            return openpyxl.utils.get_column_letter(idx)
    return None


def ensure_visit_columns(
    wb: openpyxl.workbook.workbook.Workbook,
    sheet_name: str,
    header_row: int,
) -> tuple[str, str]:
    """
    Asegura columnas de última/próxima visita. Reutiliza alias conocidos o crea 'Fecha …'.
    """
    ws = wb[sheet_name]
    headers = _header_row_values(ws, header_row)

    last_hdr = "Fecha última visita"
    next_hdr = "Fecha próxima visita"

    def find_or_append(field_key: str, default_title: str) -> str:
        letter = _col_letter_for_field(headers, field_key)
        if letter:
            return letter
        headers_lower = [h.lower() for h in headers]
        if default_title.lower() in headers_lower:
            idx = headers_lower.index(default_title.lower()) + 1
            return openpyxl.utils.get_column_letter(idx)
        max_c = ws.max_column or len(headers)
        col = max_c + 1
        ws.cell(row=header_row, column=col, value=default_title)
        headers.append(default_title)
        return openpyxl.utils.get_column_letter(col)

    col_last = find_or_append("last_visit", last_hdr)
    col_next = find_or_append("next_visit", next_hdr)
    return col_last, col_next


def write_next_visits(
    wb: openpyxl.workbook.workbook.Workbook,
    sheet_name: str,
    header_row: int,
    row_to_next: dict[int, date],
) -> None:
    """Escribe fechas de próxima visita por índice de fila Excel."""
    col_last, col_next = ensure_visit_columns(wb, sheet_name, header_row)
    ws = wb[sheet_name]
    col_next_idx = openpyxl.utils.column_index_from_string(col_next)
    for row_idx, d in row_to_next.items():
        cell = ws.cell(row=row_idx, column=col_next_idx, value=d)
        cell.number_format = "mm/dd/yyyy"


def write_last_visits(
    wb: openpyxl.workbook.workbook.Workbook,
    sheet_name: str,
    header_row: int,
    row_to_last: dict[int, date],
) -> None:
    """
    Escribe fechas de última visita y, si coincide con una programada (margen 2 semanas), 
    limpia la fecha programada.
    """
    col_last, col_next = ensure_visit_columns(wb, sheet_name, header_row)
    ws = wb[sheet_name]
    col_last_idx = openpyxl.utils.column_index_from_string(col_last)
    col_next_idx = openpyxl.utils.column_index_from_string(col_next)

    for row_idx, d in row_to_last.items():
        # 1. Escribir última visita
        cell_last = ws.cell(row=row_idx, column=col_last_idx, value=d)
        cell_last.number_format = "mm/dd/yyyy"

        # 2. Lógica de limpieza de Próxima Visita (Regla de las 2 semanas)
        # Leer valor actual de próxima visita
        raw_next = ws.cell(row=row_idx, column=col_next_idx).value
        next_date = _parse_date_cell(raw_next)
        
        if next_date:
            # Si la visita realizada está dentro de 14 días (antes o después) de la programada
            # la consideramos cumplida y limpiamos la programación.
            diff = abs((d - next_date).days)
            if diff <= 14:
                ws.cell(row=row_idx, column=col_next_idx, value=None)


def save_workbook(wb: openpyxl.workbook.workbook.Workbook, path: str | Path) -> Path:
    p = Path(path)
    wb.save(str(p))
    return p


# ── Hojas de agenda para Excel exportado ──────────────────────────

_DIAS_ES = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]


def write_agenda_flat_sheet(
    wb: openpyxl.workbook.workbook.Workbook,
    visits: list,
) -> None:
    """
    Agrega una hoja «Agenda» con listado plano de todas las visitas planificadas.
    ``visits`` es una lista de objetos con ``.visit_at`` (datetime) y
    ``.client`` (``.name``, ``.address_raw``, ``.department``).
    """
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter as gcl

    SHEET = "Agenda"
    if SHEET in wb.sheetnames:
        del wb[SHEET]
    ws = wb.create_sheet(SHEET)

    headers = ["Fecha", "Día", "Hora", "Cliente", "Dirección", "Departamento"]
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill("solid", fgColor="D31920")

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hf
        c.fill = hfill
        c.alignment = Alignment(horizontal="center")

    for idx, v in enumerate(visits, 2):
        d = v.visit_at.date()
        ws.cell(row=idx, column=1, value=d)
        ws.cell(row=idx, column=2, value=_DIAS_ES[d.weekday()])
        ws.cell(row=idx, column=3, value=v.visit_at.strftime("%H:%M"))
        ws.cell(row=idx, column=4, value=v.client.name or "")
        ws.cell(row=idx, column=5, value=v.client.address_raw or "")
        ws.cell(row=idx, column=6, value=v.client.department or "")

    for ci, w in enumerate([14, 12, 8, 38, 32, 18], 1):
        ws.column_dimensions[gcl(ci)].width = w


def write_weekly_agenda_sheet(
    wb: openpyxl.workbook.workbook.Workbook,
    visits: list,
) -> None:
    """
    Agrega una hoja «Vista Semanal» con formato de agenda:
    filas = franja horaria, columnas = lunes a viernes.
    Cada semana tiene su propio bloque con encabezado.
    """
    from collections import defaultdict
    from datetime import timedelta as _td

    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter as gcl

    SHEET = "Vista Semanal"
    if SHEET in wb.sheetnames:
        del wb[SHEET]
    ws = wb.create_sheet(SHEET)

    if not visits:
        ws.cell(row=1, column=1, value="(sin visitas)")
        return

    visits_by_date: dict[date, list] = defaultdict(list)
    for v in visits:
        visits_by_date[v.visit_at.date()].append(v)

    all_dates = sorted(visits_by_date.keys())

    # Estilos
    wh_font = Font(bold=True, size=12, color="FFFFFF")
    wh_fill = PatternFill("solid", fgColor="D31920")
    dh_font = Font(bold=True, size=10, color="FFFFFF")
    dh_fill = PatternFill("solid", fgColor="6B7B8D")
    tf = Font(bold=True, size=9, color="D31920")
    cf = Font(size=10)
    border = Border(
        left=Side("thin", color="CCCCCC"),
        right=Side("thin", color="CCCCCC"),
        top=Side("thin", color="CCCCCC"),
        bottom=Side("thin", color="CCCCCC"),
    )

    row = 1
    monday = all_dates[0] - _td(days=all_dates[0].weekday())
    last = all_dates[-1]
    DIAS = _DIAS_ES[:5]  # lun-vie

    while monday <= last:
        friday = monday + _td(days=4)

        # Encabezado de semana
        c = ws.cell(
            row=row, column=1,
            value=f"Semana del {monday.strftime('%d/%m/%Y')} "
                  f"al {friday.strftime('%d/%m/%Y')}",
        )
        c.font = wh_font
        c.fill = wh_fill
        c.alignment = Alignment(horizontal="center")
        ws.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=6,
        )
        row += 1

        # Cabecera de días
        ws.cell(row=row, column=1, value="Hora").font = dh_font
        ws.cell(row=row, column=1).fill = dh_fill
        for i, dn in enumerate(DIAS):
            dd = monday + _td(days=i)
            cell = ws.cell(
                row=row, column=i + 2,
                value=f"{dn} {dd.strftime('%d/%m')}",
            )
            cell.font = dh_font
            cell.fill = dh_fill
        row += 1

        # Slots de la semana
        slots: dict[str, dict[int, str]] = {}
        for i in range(5):
            d = monday + _td(days=i)
            for v in visits_by_date.get(d, []):
                t = v.visit_at.strftime("%H:%M")
                slots.setdefault(t, {})[i] = v.client.name or ""

        for t in sorted(slots):
            ws.cell(row=row, column=1, value=t).font = tf
            for i in range(5):
                cell = ws.cell(
                    row=row, column=i + 2,
                    value=slots[t].get(i, ""),
                )
                cell.font = cf
                cell.border = border
            row += 1

        row += 1  # fila vacía entre semanas
        monday += _td(days=7)

    # Anchos de columna
    ws.column_dimensions["A"].width = 8
    for i in range(5):
        ws.column_dimensions[gcl(i + 2)].width = 30
