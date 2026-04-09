"""Genera LISTA_DE_INDUSTRIA_ejemplo.xlsx para probar la app de escritorio."""
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "data" / "LISTA_DE_INDUSTRIA_ejemplo.xlsx"

HEADERS = [
    "N. Cli.",
    "Nombre",
    "Departamento",
    "Dirección",
    "Tipo",
    "Potencial",
    "Teléfono",
    "Oficina",
    "Vendedor",
    "Sector",
    "Fecha última visita",
    "Fecha próxima visita",
    "Latitud",
    "Longitud",
]

ROWS = [
    ["1001", "Industria Alfa", "Montevideo", "Av. Italia 1234", "grande", "alto", "24001234", "MVD", "V1", "A", "", "", -34.8941, -56.0675],
    ["1002", "Beta SRL", "Montevideo", "Bulevar Artigas 500", "medio", "medio", "", "MVD", "V1", "B", "2025-06-01", "", -34.90, -56.15],
    ["1003", "Gamma", "Canelones", "Ruta 5 km 20", "chico", "bajo", "", "CAN", "V1", "C", "", "", None, None],
    ["1004", "Delta", "Montevideo", "Ciudad Vieja 10", "chico", "alto", "", "MVD", "V2", "D", "", "", -34.907, -56.203],
]


def main() -> None:
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LISTA DE INDUSTRIA"

    hf = Font(bold=True, color="FFFFFF")
    hfil = PatternFill("solid", fgColor="2563EB")

    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = hf
        cell.fill = hfil
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = min(len(h) + 4, 40)

    for r, row in enumerate(ROWS, start=3):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)

    wb.save(OUT)
    print(f"Escrito: {OUT}")


if __name__ == "__main__":
    main()
